"""Unit tests for ExcelIntelligenceToolkit."""
import pytest
import openpyxl
from parrot.tools.excel_intelligence import ExcelIntelligenceToolkit


@pytest.fixture
def complex_excel_path(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"] = "Roadshow Net Revenue"
    ws["A2"], ws["B2"], ws["C2"] = "Client", "Jan 2024", "Feb 2024"
    ws["A3"], ws["B3"], ws["C3"] = "Client A", 10000, 12000
    ws["A4"], ws["B4"], ws["C4"] = "Client B", 8000, 9500
    ws["A5"], ws["B5"], ws["C5"] = "Total", 18000, 21500
    ws["A9"], ws["B9"], ws["C9"] = "Division", "Q1", "Q2"
    ws["A10"], ws["B10"], ws["C10"] = "North", 5000, 6200
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def toolkit():
    return ExcelIntelligenceToolkit()


class TestInspectWorkbook:
    @pytest.mark.asyncio
    async def test_returns_structural_map(self, toolkit, complex_excel_path):
        result = await toolkit.inspect_workbook(str(complex_excel_path))
        assert "Revenue" in result
        assert "T1" in result

    @pytest.mark.asyncio
    async def test_specific_sheet(self, toolkit, complex_excel_path):
        result = await toolkit.inspect_workbook(
            str(complex_excel_path), sheet_name="Revenue"
        )
        assert "Revenue" in result

    @pytest.mark.asyncio
    async def test_invalid_sheet(self, toolkit, complex_excel_path):
        result = await toolkit.inspect_workbook(
            str(complex_excel_path), sheet_name="NonExistent"
        )
        assert "not found" in result


class TestExtractTable:
    @pytest.mark.asyncio
    async def test_extract_markdown(self, toolkit, complex_excel_path):
        await toolkit.inspect_workbook(str(complex_excel_path))
        result = await toolkit.extract_table(
            str(complex_excel_path), "Revenue", "T1"
        )
        assert "Client" in result

    @pytest.mark.asyncio
    async def test_extract_csv(self, toolkit, complex_excel_path):
        await toolkit.inspect_workbook(str(complex_excel_path))
        result = await toolkit.extract_table(
            str(complex_excel_path), "Revenue", "T1", output_format="csv"
        )
        assert "Client" in result

    @pytest.mark.asyncio
    async def test_invalid_table_id(self, toolkit, complex_excel_path):
        await toolkit.inspect_workbook(str(complex_excel_path))
        result = await toolkit.extract_table(
            str(complex_excel_path), "Revenue", "T99"
        )
        assert "not found" in result


class TestQueryCells:
    @pytest.mark.asyncio
    async def test_query_range(self, toolkit, complex_excel_path):
        result = await toolkit.query_cells(
            str(complex_excel_path), "Revenue", "A1:C3"
        )
        assert "Roadshow" in result or "Client" in result


class TestToolGeneration:
    def test_tools_auto_generated(self, toolkit):
        tools = toolkit.get_tools_sync()
        tool_names = [t.name for t in tools]
        assert "inspect_workbook" in tool_names
        assert "extract_table" in tool_names
        assert "query_cells" in tool_names


class TestCleanup:
    @pytest.mark.asyncio
    async def test_cleanup(self, toolkit, complex_excel_path):
        await toolkit.inspect_workbook(str(complex_excel_path))
        assert len(toolkit._analyzer_cache) > 0
        await toolkit.cleanup()
        assert len(toolkit._analyzer_cache) == 0
