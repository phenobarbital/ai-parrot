"""Integration tests for DatasetManager file loading."""
import pytest
import openpyxl
from parrot.tools.dataset_manager.tool import DatasetManager


@pytest.fixture
def complex_excel_path(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"] = "Roadshow Net Revenue"
    ws["A2"], ws["B2"], ws["C2"] = "Client", "Jan 2024", "Feb 2024"
    ws["A3"], ws["B3"], ws["C3"] = "Client A", 10000, 12000
    ws["A4"], ws["B4"], ws["C4"] = "Client B", 8000, 9500
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def csv_path(tmp_path):
    path = tmp_path / "test.csv"
    path.write_text("Name,Age,City\nAlice,30,NYC\nBob,25,LA\n")
    return path


@pytest.fixture
def manager():
    return DatasetManager(generate_guide=False)


class TestLoadFile:
    @pytest.mark.asyncio
    async def test_load_csv(self, manager, csv_path):
        result = await manager.load_file("test_csv", csv_path)
        assert "csv" in result.lower() or "Name" in result
        assert "test_csv" in manager._file_entries

    @pytest.mark.asyncio
    async def test_load_excel(self, manager, complex_excel_path):
        result = await manager.load_file("test_excel", complex_excel_path)
        assert "Revenue" in result
        assert "test_excel" in manager._file_entries
        entry = manager._file_entries["test_excel"]
        assert entry.file_type == "excel"
        assert len(entry.markdown_content) >= 1

    @pytest.mark.asyncio
    async def test_unsupported_file(self, manager, tmp_path):
        path = tmp_path / "test.txt"
        path.write_text("hello")
        with pytest.raises(ValueError, match="Unsupported"):
            await manager.load_file("test", path)


class TestGetFileContext:
    @pytest.mark.asyncio
    async def test_get_context(self, manager, csv_path):
        await manager.load_file("test_csv", csv_path)
        context = await manager.get_file_context("test_csv")
        assert "Name" in context

    @pytest.mark.asyncio
    async def test_not_found(self, manager):
        with pytest.raises((KeyError, ValueError)):
            await manager.get_file_context("nonexistent")


class TestGetFileTable:
    @pytest.mark.asyncio
    async def test_get_table(self, manager, complex_excel_path):
        await manager.load_file("test_excel", complex_excel_path)
        entry = manager._file_entries["test_excel"]
        table_id = list(entry.markdown_content.keys())[0]
        result = await manager.get_file_table("test_excel", table_id)
        assert "Client" in result


class TestSeparation:
    @pytest.mark.asyncio
    async def test_file_and_dataframe_separate(self, manager, csv_path):
        # Load as file
        await manager.load_file("file_csv", csv_path)
        # Load as dataframe
        manager.add_dataframe_from_file("df_csv", csv_path)
        assert "file_csv" in manager._file_entries
        assert "df_csv" in manager._datasets
        assert "file_csv" not in manager._datasets
        assert "df_csv" not in manager._file_entries
