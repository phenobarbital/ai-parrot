"""Unit tests for ExcelStructureAnalyzer."""
import pytest
import openpyxl
from parrot.tools.dataset_manager.excel_analyzer import (
    ExcelStructureAnalyzer, DetectedTable, SheetAnalysis, CellRegion
)


@pytest.fixture
def complex_excel_path(tmp_path):
    """Create a test Excel file with multiple stacked tables."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"
    # Section title
    ws["A1"] = "Roadshow Net Revenue"
    # Table 1: header at row 2
    ws["A2"], ws["B2"], ws["C2"] = "Client", "Jan 2024", "Feb 2024"
    ws["A3"], ws["B3"], ws["C3"] = "Client A", 10000, 12000
    ws["A4"], ws["B4"], ws["C4"] = "Client B", 8000, 9500
    ws["A5"], ws["B5"], ws["C5"] = "Total", 18000, 21500
    # Gap rows 6-7 (empty)
    # Table 2: section title at row 8, header at row 9
    ws["A8"] = "EBITDA Summary"
    ws["A9"], ws["B9"], ws["C9"] = "Division", "Q1", "Q2"
    ws["A10"], ws["B10"], ws["C10"] = "North", 5000, 6200
    ws["A11"], ws["B11"], ws["C11"] = "South", 3200, 4100
    path = tmp_path / "test_complex.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def simple_excel_path(tmp_path):
    """Single-table Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "Name", "Age", "City"
    ws["A2"], ws["B2"], ws["C2"] = "Alice", 30, "NYC"
    ws["A3"], ws["B3"], ws["C3"] = "Bob", 25, "LA"
    path = tmp_path / "test_simple.xlsx"
    wb.save(path)
    return path


class TestCellRegion:
    def test_excel_range(self):
        region = CellRegion(start_row=2, start_col=1, end_row=5, end_col=3)
        assert region.excel_range == "A2:C5"

    def test_row_count(self):
        region = CellRegion(start_row=2, start_col=1, end_row=5, end_col=3)
        assert region.row_count == 4

    def test_col_count(self):
        region = CellRegion(start_row=2, start_col=1, end_row=5, end_col=3)
        assert region.col_count == 3


class TestExcelStructureAnalyzer:
    def test_analyze_simple_workbook(self, simple_excel_path):
        analyzer = ExcelStructureAnalyzer(simple_excel_path)
        result = analyzer.analyze_workbook()
        assert "Sheet" in result
        sheet = result["Sheet"]
        assert len(sheet.tables) >= 1
        assert sheet.tables[0].columns[0] == "Name"
        analyzer.close()

    def test_multi_table_detection(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        sheet = result["Revenue"]
        assert len(sheet.tables) >= 2
        analyzer.close()

    def test_total_row_detection(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        sheet = result["Revenue"]
        table1 = sheet.tables[0]
        assert table1.has_total_row is True
        analyzer.close()

    def test_section_title_detection(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        sheet = result["Revenue"]
        table1 = sheet.tables[0]
        assert table1.section_label is not None
        assert "Roadshow" in table1.section_label
        analyzer.close()

    def test_extract_table_as_dataframe(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        table = result["Revenue"].tables[0]
        df = analyzer.extract_table_as_dataframe("Revenue", table, include_totals=False)
        assert len(df) >= 2  # At least 2 data rows (excluding total)
        assert "Client" in df.columns or df.columns[0] == "Client"
        analyzer.close()

    def test_extract_table_with_totals(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        table = result["Revenue"].tables[0]
        df_no_totals = analyzer.extract_table_as_dataframe("Revenue", table, include_totals=False)
        df_with_totals = analyzer.extract_table_as_dataframe("Revenue", table, include_totals=True)
        assert len(df_with_totals) > len(df_no_totals)
        analyzer.close()

    def test_extract_cell_range(self, simple_excel_path):
        analyzer = ExcelStructureAnalyzer(simple_excel_path)
        rows = analyzer.extract_cell_range("Sheet", "A1:C2")
        assert len(rows) == 2
        assert rows[0][0] == "Name"
        analyzer.close()

    def test_empty_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        analyzer = ExcelStructureAnalyzer(path)
        result = analyzer.analyze_workbook()
        assert result["Empty"].total_rows == 0
        assert len(result["Empty"].tables) == 0
        analyzer.close()

    def test_to_summary(self, complex_excel_path):
        analyzer = ExcelStructureAnalyzer(complex_excel_path)
        result = analyzer.analyze_workbook()
        summary = result["Revenue"].to_summary()
        assert "Revenue" in summary
        assert "Detected tables:" in summary
        analyzer.close()
