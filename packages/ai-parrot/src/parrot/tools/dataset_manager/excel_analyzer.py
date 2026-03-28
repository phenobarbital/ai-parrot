"""Excel Structure Analysis Engine.

Scans complex Excel workbooks using openpyxl and discovers table structures
via header-row heuristics. Produces SheetAnalysis and DetectedTable data models
describing the structural layout.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)

TOTAL_KEYWORDS: set[str] = {
    "total", "sum", "subtotal", "grand total", "net total",
    "totals", "aggregate", "overall",
}

# Maximum file size before a warning is logged (100 MB).
_MAX_FILE_SIZE_BYTES = 100 * 1024 * 1024


@dataclass
class CellRegion:
    """A rectangular region within a sheet."""

    start_row: int
    start_col: int
    end_row: int
    end_col: int

    @property
    def excel_range(self) -> str:
        """Return the region as an Excel-style range string (e.g. 'A2:C5')."""
        sc = get_column_letter(self.start_col)
        ec = get_column_letter(self.end_col)
        return f"{sc}{self.start_row}:{ec}{self.end_row}"

    @property
    def row_count(self) -> int:
        """Number of rows in the region."""
        return self.end_row - self.start_row + 1

    @property
    def col_count(self) -> int:
        """Number of columns in the region."""
        return self.end_col - self.start_col + 1


@dataclass
class DetectedTable:
    """A table discovered within a sheet."""

    table_id: str
    title: Optional[str]
    header_row: int
    data_start_row: int
    data_end_row: int
    start_col: int
    end_col: int
    columns: List[str]
    row_count: int
    has_total_row: bool = False
    section_label: Optional[str] = None

    @property
    def excel_range(self) -> str:
        """Return the full range of the table (header → last data row)."""
        sc = get_column_letter(self.start_col)
        ec = get_column_letter(self.end_col)
        return f"{sc}{self.header_row}:{ec}{self.data_end_row}"

    def to_summary(self) -> str:
        """Return a human-readable summary of this table."""
        parts: list[str] = []
        label = f"  - **{self.table_id}**"
        if self.title:
            label += f": {self.title}"
        label += f" (range {self.excel_range}, {self.row_count} data rows)"
        parts.append(label)
        parts.append(f"    Columns: {', '.join(self.columns)}")
        if self.has_total_row:
            parts.append("    Has total/summary row")
        if self.section_label:
            parts.append(f"    Section: {self.section_label}")
        return "\n".join(parts)


@dataclass
class SheetAnalysis:
    """Complete structural analysis of one sheet."""

    name: str
    total_rows: int
    total_cols: int
    tables: List[DetectedTable] = field(default_factory=list)
    merged_cells: List[str] = field(default_factory=list)
    standalone_labels: List[Tuple[str, str]] = field(default_factory=list)

    def to_summary(self) -> str:
        """Return a human-readable summary of this sheet's structure."""
        parts = [
            f"## Sheet: {self.name}",
            f"  Dimensions: {self.total_rows} rows × {self.total_cols} cols",
        ]
        if self.tables:
            parts.append(f"  Detected tables: {len(self.tables)}")
            for table in self.tables:
                parts.append(table.to_summary())
        else:
            parts.append("  No tables detected.")
        if self.merged_cells:
            parts.append(f"  Merged cells: {', '.join(self.merged_cells[:10])}")
            if len(self.merged_cells) > 10:
                parts.append(f"    ... and {len(self.merged_cells) - 10} more")
        if self.standalone_labels:
            labels_str = ", ".join(
                f"{cell}: {val}" for cell, val in self.standalone_labels[:5]
            )
            parts.append(f"  Standalone labels: {labels_str}")
        return "\n".join(parts)


class ExcelStructureAnalyzer:
    """Core analysis engine for Excel workbooks.

    Uses ``openpyxl`` to scan sheets and discover table structures via
    header-row heuristics (3+ non-empty cells with 40 %+ strings and
    numeric data below).

    Args:
        path: Path to the Excel file.
    """

    def __init__(self, path: Union[str, Path]) -> None:
        self._path = Path(path)
        if self._path.stat().st_size > _MAX_FILE_SIZE_BYTES:
            logger.warning(
                "File '%s' is %.1f MB — analysis may be slow.",
                self._path.name,
                self._path.stat().st_size / (1024 * 1024),
            )
        # Read-only workbook for initial analysis (memory-efficient).
        self._wb_ro = openpyxl.load_workbook(
            self._path, read_only=True, data_only=True,
        )
        # Normal-mode workbook for merged cell detection and random access.
        self._wb: Optional[openpyxl.Workbook] = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def analyze_workbook(self) -> Dict[str, SheetAnalysis]:
        """Analyze all sheets and return a mapping of sheet name → SheetAnalysis."""
        results: Dict[str, SheetAnalysis] = {}
        for sheet_name in self._wb_ro.sheetnames:
            results[sheet_name] = self._analyze_sheet(sheet_name)
        return results

    def extract_table_as_dataframe(
        self,
        sheet_name: str,
        table: DetectedTable,
        include_totals: bool = True,
    ) -> pd.DataFrame:
        """Extract a detected table as a clean pandas DataFrame.

        Args:
            sheet_name: Sheet containing the table.
            table: DetectedTable instance (from analyze_workbook).
            include_totals: If False, exclude the total row (if detected).

        Returns:
            pandas DataFrame with proper column headers.
        """
        ws = self._get_normal_ws(sheet_name)
        end_row = table.data_end_row
        if not include_totals and table.has_total_row:
            end_row -= 1

        rows: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=table.data_start_row,
            max_row=end_row,
            min_col=table.start_col,
            max_col=table.end_col,
            values_only=True,
        ):
            rows.append(list(row))

        return pd.DataFrame(rows, columns=table.columns)

    def extract_cell_range(
        self, sheet_name: str, cell_range: str,
    ) -> List[List[Any]]:
        """Read raw cell values from an arbitrary Excel-style range.

        Args:
            sheet_name: Name of the sheet.
            cell_range: Excel range string (e.g. ``'A1:C5'``).

        Returns:
            List of rows, each a list of cell values.
        """
        ws = self._get_normal_ws(sheet_name)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        rows: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        ):
            rows.append(list(row))
        return rows

    def close(self) -> None:
        """Close all open workbooks."""
        try:
            self._wb_ro.close()
        except Exception:
            pass
        if self._wb is not None:
            try:
                self._wb.close()
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_normal_ws(self, sheet_name: str):
        """Return a worksheet from the normal-mode workbook (lazy-loaded)."""
        if self._wb is None:
            self._wb = openpyxl.load_workbook(
                self._path, read_only=False, data_only=True,
            )
        return self._wb[sheet_name]

    def _analyze_sheet(self, sheet_name: str) -> SheetAnalysis:
        """Analyze a single sheet for tables and standalone labels."""
        ws_ro = self._wb_ro[sheet_name]

        # Materialise all rows for random access during analysis.
        all_rows: list[tuple] = []
        for row in ws_ro.iter_rows(values_only=True):
            all_rows.append(tuple(row))

        total_rows = len(all_rows)
        total_cols = max((len(r) for r in all_rows), default=0) if all_rows else 0

        # Detect merged cells (need normal-mode workbook).
        merged_cells: list[str] = []
        try:
            ws_normal = self._get_normal_ws(sheet_name)
            merged_cells = [str(mc) for mc in ws_normal.merged_cells.ranges]
        except Exception:
            pass

        # Discover tables.
        tables: list[DetectedTable] = []
        standalone_labels: list[tuple[str, str]] = []
        table_counter = 0
        consumed_rows: set[int] = set()  # rows already part of a table

        for row_idx in range(total_rows):
            if row_idx in consumed_rows:
                continue
            if self._is_header_row(all_rows, row_idx, total_cols):
                table_counter += 1
                table = self._extract_table_from_header(
                    all_rows, row_idx, total_rows, total_cols, table_counter,
                )
                if table is not None:
                    tables.append(table)
                    for r in range(table.header_row - 1, table.data_end_row):
                        consumed_rows.add(r)

        # Identify standalone labels (non-empty cells not part of any table).
        for row_idx in range(total_rows):
            if row_idx in consumed_rows:
                continue
            row = all_rows[row_idx]
            non_empty = [(ci, v) for ci, v in enumerate(row) if v is not None]
            if len(non_empty) == 1:
                ci, val = non_empty[0]
                if isinstance(val, str) and val.strip():
                    cell_ref = f"{get_column_letter(ci + 1)}{row_idx + 1}"
                    standalone_labels.append((cell_ref, val.strip()))

        return SheetAnalysis(
            name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            tables=tables,
            merged_cells=merged_cells,
            standalone_labels=standalone_labels,
        )

    # ------------------------------------------------------------------
    # Heuristics
    # ------------------------------------------------------------------

    def _is_header_row(
        self, all_rows: list[tuple], row_idx: int, total_cols: int,
    ) -> bool:
        """Return True if *row_idx* looks like a table header row.

        A header row has 3+ non-empty cells where 40 %+ are strings, and the
        row immediately below contains at least one numeric/date value.
        """
        if row_idx >= len(all_rows) - 1:
            return False  # need at least one data row below

        row = all_rows[row_idx]
        non_empty = [v for v in row if v is not None]
        if len(non_empty) < 3:
            return False

        # Check that ≥40 % of non-empty cells are strings.
        str_count = sum(1 for v in non_empty if isinstance(v, str))
        if str_count / len(non_empty) < 0.4:
            return False

        # Check that the row below has at least one numeric or date value.
        next_row = all_rows[row_idx + 1]
        has_numeric = any(
            isinstance(v, (int, float, datetime))
            for v in next_row if v is not None
        )
        return has_numeric

    def _extract_table_from_header(
        self,
        all_rows: list[tuple],
        header_idx: int,
        total_rows: int,
        total_cols: int,
        table_number: int,
    ) -> Optional[DetectedTable]:
        """Expand a header row into a full DetectedTable.

        * Scans downward until 2 consecutive empty rows.
        * Detects total rows by keyword matching.
        * Looks 1-3 rows above the header for a section title.
        """
        header = all_rows[header_idx]

        # Determine column span (first non-empty to last non-empty in header).
        col_indices = [i for i, v in enumerate(header) if v is not None]
        if not col_indices:
            return None
        start_col = col_indices[0]
        end_col = col_indices[-1]

        columns: list[str] = []
        for ci in range(start_col, end_col + 1):
            val = header[ci] if ci < len(header) else None
            if val is None:
                columns.append(f"col_{ci + 1}")
            elif isinstance(val, datetime):
                columns.append(val.strftime("%b %Y"))
            else:
                columns.append(str(val).strip())

        # Expand downward — stop at 2 consecutive empty rows.
        data_start = header_idx + 1
        data_end = data_start  # inclusive
        consecutive_empty = 0

        for ri in range(data_start, total_rows):
            row = all_rows[ri]
            non_empty = [
                row[ci] for ci in range(start_col, min(end_col + 1, len(row)))
                if ci < len(row) and row[ci] is not None
            ]
            if not non_empty:
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
            else:
                consecutive_empty = 0
                data_end = ri

        if data_end < data_start:
            return None

        # Detect total row (last data row whose first cell matches a keyword).
        has_total = False
        last_row = all_rows[data_end]
        first_cell = (
            last_row[start_col]
            if start_col < len(last_row) else None
        )
        if first_cell is not None and isinstance(first_cell, str):
            normalised = first_cell.strip().lower()
            if normalised in TOTAL_KEYWORDS or any(
                kw in normalised for kw in TOTAL_KEYWORDS
            ):
                has_total = True

        # Look 1-3 rows above header for section title.
        section_label: Optional[str] = None
        for lookback in range(1, 4):
            ri = header_idx - lookback
            if ri < 0:
                break
            row = all_rows[ri]
            non_empty = [v for v in row if v is not None]
            if len(non_empty) == 1 and isinstance(non_empty[0], str):
                section_label = non_empty[0].strip()
                break

        row_count = data_end - data_start + 1
        table_id = f"T{table_number}"
        title = section_label or columns[0] if columns else None

        return DetectedTable(
            table_id=table_id,
            title=title,
            header_row=header_idx + 1,  # 1-based
            data_start_row=data_start + 1,  # 1-based
            data_end_row=data_end + 1,  # 1-based
            start_col=start_col + 1,  # 1-based
            end_col=end_col + 1,  # 1-based
            columns=columns,
            row_count=row_count,
            has_total_row=has_total,
            section_label=section_label,
        )
