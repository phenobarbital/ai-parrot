"""Unit tests for ExcelLoader sheet mode and row mode.

Covers all test cases from FEAT-093 spec Section 4:
- Sheet mode: one Document per non-empty sheet with structural context
- Row mode: one Document per row (legacy backward compatibility)
- DataFrame input: falls back to row mode regardless of output_mode
"""
import pytest
import openpyxl
import pandas as pd

from parrot_loaders.excel import ExcelLoader
from parrot.stores.models import Document


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def simple_excel(tmp_path):
    """Single-sheet workbook with 5 data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Product", "Revenue", "Units"])
    for i in range(5):
        ws.append([f"Product {i}", (i + 1) * 100, (i + 1) * 10])
    path = tmp_path / "simple.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_excel(tmp_path):
    """3-sheet workbook: Sales (3 rows), Expenses (2 rows), Empty."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Q1", "Q2"])
    for i in range(3):
        ws1.append([f"Item {i}", i * 10, i * 20])
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 5000])
    ws2.append(["Utilities", 800])
    wb.create_sheet("Empty")
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def large_table_excel(tmp_path):
    """Single-sheet workbook with 50 data rows for truncation testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Value", "Category"])
    for i in range(50):
        ws.append([i, i * 3.14, f"Cat_{i % 5}"])
    path = tmp_path / "large.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Sheet Mode Tests
# ---------------------------------------------------------------------------

class TestExcelLoaderSheetMode:
    """Tests for output_mode='sheet' (new per-sheet document generation)."""

    @pytest.mark.asyncio
    async def test_sheet_mode_one_doc_per_sheet(self, simple_excel):
        """Single-sheet workbook -> exactly 1 Document."""
        loader = ExcelLoader(source=str(simple_excel), output_mode="sheet")
        docs = await loader.load()
        assert len(docs) == 1
        assert isinstance(docs[0], Document)

    @pytest.mark.asyncio
    async def test_sheet_mode_multi_sheet(self, multi_sheet_excel):
        """3-sheet workbook (1 empty) -> exactly 2 Documents."""
        loader = ExcelLoader(source=str(multi_sheet_excel), output_mode="sheet")
        docs = await loader.load()
        assert len(docs) == 2
        sheet_names = [
            d.metadata["document_meta"]["sheet"] for d in docs
        ]
        assert "Sales" in sheet_names
        assert "Expenses" in sheet_names
        assert "Empty" not in sheet_names

    @pytest.mark.asyncio
    async def test_sheet_mode_structural_header(self, simple_excel):
        """Document contains sheet name, dimensions, table summary."""
        loader = ExcelLoader(source=str(simple_excel), output_mode="sheet")
        docs = await loader.load()
        content = docs[0].page_content
        assert "Sheet: Sales" in content
        assert "Dimensions:" in content
        assert "Detected tables:" in content

    @pytest.mark.asyncio
    async def test_sheet_mode_tables_as_markdown(self, simple_excel):
        """Detected tables rendered as markdown with | separators."""
        loader = ExcelLoader(source=str(simple_excel), output_mode="sheet")
        docs = await loader.load()
        content = docs[0].page_content
        # Markdown tables use | as column separator
        assert "|" in content
        assert "Product" in content
        assert "Revenue" in content

    @pytest.mark.asyncio
    async def test_sheet_mode_metadata(self, simple_excel):
        """Metadata has content_type: 'sheet', table_count, sheet name."""
        loader = ExcelLoader(source=str(simple_excel), output_mode="sheet")
        docs = await loader.load()
        meta = docs[0].metadata["document_meta"]
        assert meta["content_type"] == "sheet"
        assert "table_count" in meta
        assert meta["table_count"] >= 0
        assert meta["sheet"] == "Sales"
        assert "tables" in meta
        assert isinstance(meta["tables"], list)

    @pytest.mark.asyncio
    async def test_sheet_mode_max_rows_truncation(self, large_table_excel):
        """Tables exceeding max_rows_per_table are truncated."""
        loader = ExcelLoader(
            source=str(large_table_excel),
            output_mode="sheet",
            max_rows_per_table=10,
        )
        docs = await loader.load()
        assert len(docs) == 1
        content = docs[0].page_content
        # Should have truncation note
        assert "truncated" in content.lower()
        # Count data rows in the markdown table (lines starting with |
        # that are not header separators)
        lines = content.split("\n")
        table_data_lines = [
            line for line in lines
            if line.startswith("|") and "---" not in line
        ]
        # Should have at most 10 data rows + 1 header row = 11 lines
        assert len(table_data_lines) <= 11

    @pytest.mark.asyncio
    async def test_sheet_mode_empty_sheet_skipped(self, multi_sheet_excel):
        """Empty sheets produce no Documents."""
        loader = ExcelLoader(source=str(multi_sheet_excel), output_mode="sheet")
        docs = await loader.load()
        sheet_names = [
            d.metadata["document_meta"]["sheet"] for d in docs
        ]
        assert "Empty" not in sheet_names

    @pytest.mark.asyncio
    async def test_sheet_mode_no_tables_raw_content(self, tmp_path):
        """Sheets with no detected tables produce Document with raw cell content."""
        # Create a workbook with text-only content (no numeric data below header)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notes"
        ws.append(["Note A"])
        ws.append(["Note B"])
        ws.append(["Note C"])
        path = tmp_path / "text_only.xlsx"
        wb.save(path)

        loader = ExcelLoader(source=str(path), output_mode="sheet")
        docs = await loader.load()
        assert len(docs) == 1
        content = docs[0].page_content
        # Should contain raw cell content since no tables detected
        assert "Note" in content


# ---------------------------------------------------------------------------
# Row Mode Tests
# ---------------------------------------------------------------------------

class TestExcelLoaderRowMode:
    """Tests for output_mode='row' (legacy backward compatibility)."""

    @pytest.mark.asyncio
    async def test_row_mode_backward_compat(self, simple_excel):
        """output_mode='row' produces per-row Documents."""
        loader = ExcelLoader(source=str(simple_excel), output_mode="row")
        docs = await loader.load()
        # 5 data rows
        assert len(docs) == 5
        # Each doc should have row-level metadata
        for doc in docs:
            meta = doc.metadata["document_meta"]
            assert meta["content_type"] == "row"
            assert "row_index" in meta

    @pytest.mark.asyncio
    async def test_default_mode_is_sheet(self):
        """Default output_mode is 'sheet'."""
        loader = ExcelLoader()
        assert loader.output_mode == "sheet"


# ---------------------------------------------------------------------------
# DataFrame Input Tests
# ---------------------------------------------------------------------------

class TestExcelLoaderDataFrameInput:
    """Tests for DataFrame input (always falls back to row mode)."""

    @pytest.mark.asyncio
    async def test_dataframe_input_falls_back_to_row(self):
        """DataFrame input uses row mode regardless of output_mode."""
        df = pd.DataFrame({
            "Name": ["Alice", "Bob", "Charlie"],
            "Score": [95, 87, 72],
        })
        loader = ExcelLoader(output_mode="sheet")
        docs = await loader.load(source=df)
        # Should produce per-row documents (3 rows)
        assert len(docs) == 3
        for doc in docs:
            meta = doc.metadata["document_meta"]
            assert meta["content_type"] == "row"
            assert meta["sheet"] == "DataFrame"
